from flask import Flask, request, render_template, send_from_directory, abort 
import fitz  # PyMuPDF
import os
import re
import pandas as pd
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation

app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
PREVIEW_FOLDER = 'static/previews'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['PREVIEW_FOLDER'] = PREVIEW_FOLDER

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(PREVIEW_FOLDER, exist_ok=True)

DIMENSION_PATTERN = re.compile(
    r"(⌀?\d+(?:[\.,]\d+)?(?:°)?)"
    r"(?:[ ]?[\u00b1]?[ ]?(\d+(?:[\.,]\d+)?))?",
    re.VERBOSE
)

NEG_TOLERANCE_PATTERN = re.compile(r"-\d+[\.,]?\d*")
PLUS_MINUS_PATTERN = re.compile(r"±\s*(\d+[\.,]?\d*)")
TAPPED_HOLE_PATTERN = re.compile(r"^(M\d+)\b")

TOLERANCE_RANGES = [
    (0.1, 6, 0.1),
    (6, 30, 0.2),
    (30, 120, 0.3),
    (120, 315, 0.5),
    (315, 1000, 0.8),
    (1000, 1200, 1.2)
]

MEASURING_INSTRUMENTS = ["Vernier", "Mic", "HG", "BG", "CMM", "SG", "PG", "TG"]

def calculate_general_tolerance(value):
    try:
        value = float(value)
        for lower, upper, tol in TOLERANCE_RANGES:
            if lower <= value < upper:
                return tol
    except:
        pass
    return None

def is_in_working_area(y, page_height):
    return (page_height * 0.05) < y < (page_height * 0.82)

def highlight_and_balloon(page, text, position, balloon_counter):
    try:
        highlight_rects = page.search_for(text)
        for rect in highlight_rects:
            page.add_highlight_annot(rect)

        radius = 10
        font_size = 10
        x, y = position
        balloon_x = x + len(text) / 2 - radius
        balloon_y = y - radius - 2

        page_rect = page.rect
        balloon_x = max(balloon_x, page_rect.x0 + 5)
        balloon_y = max(balloon_y, page_rect.y0 + 5)

        balloon_rect = fitz.Rect(
            balloon_x - radius,
            balloon_y - radius,
            balloon_x + radius,
            balloon_y + radius
        )
        page.draw_oval(balloon_rect, color=(1, 0, 0), fill=(1, 1, 1), width=1.5)
        text_x = balloon_x - (len(str(balloon_counter)) * font_size * 0.25)
        text_y = balloon_y + radius - font_size * 0.75
        page.insert_text(
            (text_x, text_y),
            str(balloon_counter),
            fontsize=font_size,
            fontname="helv",
            color=(1, 0, 0)
        )
    except Exception as e:
        print(f"Balloon error: {e}")

def write_to_inspection_template(df, excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inspection Report"

    headers = [
        ["", "", "", "", "INSIGHT TECHNOLOGIES", "", "", "", "", "", ""],
        ["", "", "", "", "FINAL INSPECTION REPORT", "", "", "", "", "", ""],
        ["Part No:", df["Part No"].iloc[0] if "Part No" in df.columns else "", "", "Project No:", "", "", "Date:", ""],
        ["Part Name:", df["Part Name"].iloc[0] if "Part Name" in df.columns else "", "", "Customer Name:", "", "", "Doc No:", "", "", "Rev Date:", ""],
        ["", "", "", "", "PHYSICAL INSPECTION", "", "", "", "", "", ""],
        ["Sl No", "Characteristics", "Method", "Acceptance Criteria", "Characteristics", "Method", "Acceptance Criteria", "Remark", ""],
        ["1", "Finish", "", "", "Quantity", "", "", "", "", ""],
        ["2", "Coating", df.get("Surface Coating", [""])[0], "", "Hardness", df.get("Heat Treatment", [""])[0], "", "", "", ""],
        ["3", "Rust", "", "", "Material", df.get("Material", [""])[0], "", "", "", ""],
        ["", "", "", "", "DIMENSIONAL INSPECTION", "", "", "", "", "", ""],
        ["S.No", "Balloon No", "Characteristics", "Nominal Size", "Tol", "Upper Limit", "Lower Limit", "Measuring Instrument", "Method of inspection", "", "", "", "Number of parts", ""],
        ["", "", "", "", "", "", "", "", "", "1", "2", "3", "4", "5", "6", "7", "8", "Remark", ""]
    ]

    for r_idx, row in enumerate(headers, start=1):
        ws.append(row)
        for c_idx, cell in enumerate(row, start=1):
            if isinstance(cell, str) and cell.strip():
                ws.cell(row=r_idx, column=c_idx).font = Font(bold=True)

    start_row = ws.max_row + 1

    for idx, row in df.iterrows():
        ws.append([
            idx + 1,
            row.get("Balloon Number", ""),
            row.get("Dimension Type", ""),
            row.get("Nominal Dimension", ""),
            row.get("Tolerance", ""),
            row.get("Upper Limit", ""),
            row.get("Lower Limit", ""),
            "",  # Measuring Instrument (dropdown added below)
            "", "", "", "", "", "", "", "", "", "", ""
        ])

    dropdown_list = ",".join(MEASURING_INSTRUMENTS)
    dv = DataValidation(type="list", formula1=f'"{dropdown_list}"', allow_blank=True)

    ws.add_data_validation(dv)
    for row in range(start_row, start_row + len(df)):
        dv.add(ws.cell(row=row, column=8))  # Measuring Instrument column

    # Center align all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add and bold the final two rows
    final_rows = [
        ["Accepted By:", "", "", "Rework:", "", "", "Rejected:", ""],
        ["Checked by:", "", "", "", "", "", "Inspected by:", ""]
    ]
    for final_row in final_rows:
        ws.append(final_row)
        current_row = ws.max_row
        for col in range(1, len(final_row) + 1):
            ws.cell(row=current_row, column=col).font = Font(bold=True)

    wb.save(excel_path)


def process_pdf(filepath, filename):
    output_pdf_path = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    excel_output_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{os.path.splitext(filename)[0]}_Dimensions.xlsx")
    preview_image_path = os.path.join(app.config['PREVIEW_FOLDER'], f"{os.path.splitext(filename)[0]}.png")

    doc = fitz.open(filepath)
    results = []
    balloon_counter = 1
    part_no = ""
    part_name = ""
    surface_coating = ""
    material = ""
    heat_treatment = ""

    for page_number, page in enumerate(doc):
        if page_number == 0:
            pix = page.get_pixmap(dpi=150)
            pix.save(preview_image_path)

        text = page.get_text()

        coating_match = re.search(r"Surface Coating\s*:\s*(.*)", text, re.IGNORECASE)
        if coating_match:
            surface_coating = coating_match.group(1).split('|')[0].strip()

        material_match = re.search(r"Material\s*:\s*(.*)", text, re.IGNORECASE)
        if material_match:
            material = material_match.group(1).split('|')[0].strip()

        heat_match = re.search(r"Heat Treatment\s*:\s*(.*)", text, re.IGNORECASE)
        if heat_match:
            heat_treatment = heat_match.group(1).split('|')[0].strip()

        pn_match = re.search(r"(?:P/N|Part Number)[:\s]*([0-9]{4}-[0-9]{6})", text, re.IGNORECASE)
        if pn_match:
            part_no = pn_match.group(1)

        title_match = re.search(r"Title[:\s]*([\w\s\-/]+)", text, re.IGNORECASE)
        if title_match:
            part_name = title_match.group(1).strip()

        page_height = page.rect.height
        words = page.get_text("words", sort=True)
        processed_rects = []

        for word in words:
            x, y = word[0], word[1]
            text = word[4].strip()

            if not is_in_working_area(y, page_height):
                continue

            rect = fitz.Rect(word[:4])
            if any(rect.intersects(prev) for prev in processed_rects):
                continue

            tapped_match = TAPPED_HOLE_PATTERN.match(text)
            if tapped_match:
                tapped = tapped_match.group(1)
                highlight_and_balloon(page, tapped, (x, y), balloon_counter)
                results.append({
                    "File Name": filename,
                    "Page": page_number + 1,
                    "Dimension Type": "Tapped Hole",
                    "Balloon Number": balloon_counter,
                    "Nominal Dimension": tapped,
                    "Tolerance": "-",
                    "Upper Limit": "-",
                    "Lower Limit": "-",
                    "Part No": part_no,
                    "Part Name": part_name
                })
                balloon_counter += 1
                processed_rects.append(rect)
                continue

            match = DIMENSION_PATTERN.match(text)
            if match:
                nominal = match.group(1).replace(",", ".")
                tolerance_str = match.group(2)
                dimension_type = "Diametrical" if "⌀" in nominal else "Angular" if "°" in nominal else "Linear"
                nominal_val = float(re.sub(r"[^\d.]", "", nominal))

                nearby_texts = [
                    w[4].strip() for w in words
                    if abs(w[0] - x) < 50 and abs(w[1] - y) < 20 and w[4].strip() != text
                ]

                neg_tols = [float(t.replace(",", ".")) for nt in nearby_texts for t in NEG_TOLERANCE_PATTERN.findall(nt)]
                plus_minus_tols = [float(t.replace(",", ".")) for nt in nearby_texts for t in PLUS_MINUS_PATTERN.findall(nt)]

                if neg_tols:
                    upper = nominal_val + max(neg_tols)
                    lower = nominal_val + min(neg_tols)
                    tol_display = f"{min(neg_tols)} to {max(neg_tols)}"
                elif plus_minus_tols:
                    tol_val = plus_minus_tols[0]
                    upper = nominal_val + tol_val
                    lower = nominal_val - tol_val
                    tol_display = f"±{tol_val:.2f}"
                elif tolerance_str:
                    tol_val = float(tolerance_str.replace(",", "."))
                    upper = nominal_val + tol_val
                    lower = nominal_val - tol_val
                    tol_display = f"±{tol_val:.2f}"
                else:
                    gen_tol = calculate_general_tolerance(nominal_val)
                    if gen_tol:
                        upper = nominal_val + gen_tol
                        lower = nominal_val - gen_tol
                        tol_display = f"±{gen_tol:.2f}"
                    else:
                        upper = lower = "-"
                        tol_display = "-"

                highlight_and_balloon(page, nominal, (x, y), balloon_counter)
                results.append({
                    "File Name": filename,
                    "Page": page_number + 1,
                    "Dimension Type": dimension_type,
                    "Balloon Number": balloon_counter,
                    "Nominal Dimension": nominal,
                    "Tolerance": tol_display,
                    "Upper Limit": upper,
                    "Lower Limit": lower,
                    "Part No": part_no,
                    "Part Name": part_name
                })
                balloon_counter += 1
                processed_rects.append(rect)

    if results:
        doc.save(output_pdf_path)
        df = pd.DataFrame(results)
        df.insert(0, "Surface Coating", surface_coating)
        df.insert(1, "Material", material)
        df.insert(2, "Heat Treatment", heat_treatment)
        write_to_inspection_template(df, excel_output_path)
        return output_pdf_path, excel_output_path
    return None, None

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return render_template('index.html', error='No file part')
        files = request.files.getlist('file')
        if not files or all(file.filename == '' for file in files):
            return render_template('index.html', error='No selected files')

        results = []
        for file in files:
            if file and file.filename.lower().endswith('.pdf'):
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                pdf_path, excel_path = process_pdf(filepath, filename)
                if pdf_path and excel_path:
                    results.append({
                        'pdf_filename': os.path.basename(pdf_path),
                        'excel_filename': os.path.basename(excel_path),
                        'message': f"✅ Processed {filename}"
                    })
                else:
                    results.append({
                        'pdf_filename': None,
                        'excel_filename': None,
                        'message': f"⚠️ No dimensions detected in {filename}"
                    })

        if not results:
            return render_template('index.html', error='No valid PDF files uploaded')
        return render_template('index.html', results=results)

    return render_template('index.html')

@app.route('/download/<filename>')
def download_file(filename):
    try:
        return send_from_directory(app.config['OUTPUT_FOLDER'], filename, as_attachment=True)
    except FileNotFoundError:
        abort(404)

if __name__ == '__main__':
    app.run(debug=True)

